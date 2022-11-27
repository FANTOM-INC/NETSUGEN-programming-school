import openpyxl

# 読み込み
wb = openpyxl.load_workbook('./sample_file/sample.xlsx')

wb.create_sheet('new_sheet')  # シート追加
wb.remove_sheet(wb["new_sheet"])  # シート削除

# 既存のシートを指定
sheet = wb['Sheet1']
value = sheet.cell(row=1, column=1).value  # A1が格納される
print("A1:", value)
value = sheet['A1'].value  # この書き方でも同じ
print("A1:", value)

sheet.cell(row=1, column=1).value = "Hello"  # A1にHelloを格納

# すべてのシートを検索してHeを含むセルを取得
sheets = wb.sheetnames
for sheet_name in sheets:
    sheet = wb[sheet_name]
    for r in sheet.iter_rows():
        for c in r:
            if "He" in str(c.value):
                print(c.row)
# 保存
wb.save('./new_excel.xlsx')

# 終了
wb.close()
